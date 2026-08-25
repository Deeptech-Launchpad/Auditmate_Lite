"""Excel workbook the customer edits.

The review round-trip works because the customer gets something they already
know how to use: a spreadsheet. They type revised figures into a dedicated
column, reply to the email, and the app reads that same workbook back.

Two functions matter:
    build_workbook()   statements  ->  .xlsx the customer edits
    read_workbook()    edited .xlsx ->  {line_key: revised amount} + comments

The layout is deliberately simple. "Amount" is what we sent (locked visually
by shading), "Revised Amount" is theirs to change, "Comment" is free text.
Only rows where the revised value differs are treated as changes, so an
untouched workbook produces no false edits.
"""
import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# Column layout. Keep these in sync with read_workbook().
COL_KEY = 1        # hidden: line_key, so we can match rows on the way back
COL_LABEL = 2
COL_AMOUNT = 3
COL_REVISED = 4
COL_COMMENT = 5

HEADER_FILL = PatternFill("solid", fgColor="16233A")
GROUP_FILL = PatternFill("solid", fgColor="EEF1F5")
LOCKED_FILL = PatternFill("solid", fgColor="F7F8FA")
EDIT_FILL = PatternFill("solid", fgColor="FFF9E6")
THIN = Side(style="thin", color="D0D5DD")


def _money_format():
    return '#,##0.00;(#,##0.00);"-"'


def build_workbook(financial_year, statements, out_path: Path,
                   version_no: int, token: str) -> Path:
    """Write the customer-facing workbook. Returns the path written."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    # --- Instructions sheet -------------------------------------------------
    intro = workbook.create_sheet("Start Here")
    intro.column_dimensions["A"].width = 96

    customer = financial_year.customer
    lines = [
        ("title", f"{customer.legal_name or customer.name}"),
        ("sub", f"Financial statements for the year ended "
                f"{financial_year.end_date.strftime('%d %B %Y') if financial_year.end_date else ''}"),
        ("sub", f"Version {version_no}   |   Reference {token}"),
        ("blank", ""),
        ("head", "What we need from you"),
        ("body", "1. Review each statement on the tabs of this workbook."),
        ("body", "2. If a figure is wrong, type the correct number in the "
                 "'Revised Amount' column. Leave it blank if the figure is right."),
        ("body", "3. Add a note in the 'Comment' column if it helps explain "
                 "the change."),
        ("body", "4. Save this file and REPLY to our email with it attached."),
        ("blank", ""),
        ("warn", "Please keep the email subject line unchanged - it contains "
                 f"the reference {token} that files your reply against the "
                 "correct engagement."),
        ("blank", ""),
        ("body", "Do not add, delete or reorder rows. Only the 'Revised "
                 "Amount' and 'Comment' columns are read back."),
    ]

    row = 1
    for kind, text in lines:
        cell = intro.cell(row=row, column=1, value=text)
        if kind == "title":
            cell.font = Font(size=15, bold=True, color="16233A")
        elif kind == "sub":
            cell.font = Font(size=10, color="6B7280")
        elif kind == "head":
            cell.font = Font(size=12, bold=True, color="1E4D8C")
        elif kind == "warn":
            cell.font = Font(size=10, bold=True, color="A86A00")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            intro.row_dimensions[row].height = 32
        else:
            cell.font = Font(size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    # --- One sheet per statement -------------------------------------------
    for statement in statements:
        title = statement.type_label[:31]
        sheet = workbook.create_sheet(title)

        sheet.column_dimensions[get_column_letter(COL_KEY)].width = 2
        sheet.column_dimensions[get_column_letter(COL_KEY)].hidden = True
        sheet.column_dimensions[get_column_letter(COL_LABEL)].width = 46
        sheet.column_dimensions[get_column_letter(COL_AMOUNT)].width = 18
        sheet.column_dimensions[get_column_letter(COL_REVISED)].width = 18
        sheet.column_dimensions[get_column_letter(COL_COMMENT)].width = 42

        headers = ["", "Description", "Amount", "Revised Amount", "Comment"]
        for col, text in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=text)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="right" if col in
                                       (COL_AMOUNT, COL_REVISED) else "left")

        sheet.freeze_panes = "B2"
        excel_row = 2
        last_group = None

        for line in statement.lines:
            # Group heading row, so the sheet reads like the statement.
            if line.group_key and line.group_key != last_group:
                last_group = line.group_key
                cell = sheet.cell(row=excel_row, column=COL_LABEL,
                                  value=line.group_key.replace("_", " ").title())
                cell.font = Font(bold=True, size=9, color="6B7280")
                for col in range(COL_LABEL, COL_COMMENT + 1):
                    sheet.cell(row=excel_row, column=col).fill = GROUP_FILL
                excel_row += 1

            sheet.cell(row=excel_row, column=COL_KEY, value=line.line_key)

            label_cell = sheet.cell(row=excel_row, column=COL_LABEL,
                                    value=("    " if line.indent else "") + line.label)
            if line.is_total or line.is_subtotal:
                label_cell.font = Font(bold=True)

            amount_cell = sheet.cell(row=excel_row, column=COL_AMOUNT,
                                     value=float(line.effective_amount or 0))
            amount_cell.number_format = _money_format()
            amount_cell.fill = LOCKED_FILL
            if line.is_total or line.is_subtotal:
                amount_cell.font = Font(bold=True)

            revised_cell = sheet.cell(row=excel_row, column=COL_REVISED)
            revised_cell.number_format = _money_format()
            revised_cell.fill = EDIT_FILL

            for col in range(COL_LABEL, COL_COMMENT + 1):
                sheet.cell(row=excel_row, column=col).border = Border(
                    bottom=THIN)

            excel_row += 1

    workbook.save(out_path)
    log.info("Wrote customer workbook %s", out_path)
    return out_path


def _to_decimal(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).replace(",", "").strip()).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def read_workbook(path: Path) -> dict:
    """Read a customer-edited workbook back.

    Returns {"changes": [...], "comments": [...]}, where each change is
    {sheet, line_key, label, original, revised, comment}. Only rows whose
    revised value actually differs are reported, so an untouched workbook
    yields nothing.
    """
    result = {"changes": [], "comments": [], "error": None}

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:                       # noqa: BLE001
        result["error"] = f"Could not open workbook: {exc}"
        return result

    for sheet in workbook.worksheets:
        if sheet.title == "Start Here":
            continue

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < COL_REVISED:
                continue

            line_key = row[COL_KEY - 1]
            if not line_key:
                continue

            label = row[COL_LABEL - 1]
            original = _to_decimal(row[COL_AMOUNT - 1])
            revised = _to_decimal(row[COL_REVISED - 1])
            comment = row[COL_COMMENT - 1] if len(row) >= COL_COMMENT else None

            if comment and str(comment).strip():
                result["comments"].append({
                    "sheet": sheet.title,
                    "line_key": str(line_key),
                    "label": str(label or "").strip(),
                    "comment": str(comment).strip(),
                })

            # No entry, or the same number typed back = not a change.
            if revised is None:
                continue
            if original is not None and revised == original:
                continue

            result["changes"].append({
                "sheet": sheet.title,
                "line_key": str(line_key),
                "label": str(label or "").strip(),
                "original": float(original) if original is not None else None,
                "revised": float(revised),
                "comment": str(comment).strip() if comment else None,
            })

    return result
